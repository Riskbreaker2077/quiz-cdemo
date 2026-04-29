#!/data/workspace/.venv/bin/python
import json
from collections import defaultdict
from pathlib import Path
import openpyxl

WORKBOOK = Path('/data/workspace/colegio_notas.xlsx')
OUT = Path('/data/workspace/quiz-cdemo/data/eva_10a.json')

PRUEBA_META = {
    'S-I-M1': {'label': 'C. Sociales · Periodo I · Módulo 1', 'maximo': 20},
    'S-I-M2': {'label': 'C. Sociales · Periodo I · Módulo 2', 'maximo': 20},
    'F-I-M1-1': {'label': 'Filosofía · Periodo I · Módulo 1 · Prueba 1', 'maximo': 10},
    'F-I-M1-2': {'label': 'Filosofía · Periodo I · Módulo 1 · Prueba 2', 'maximo': 10},
}

ORDER = {
    'S-I-M1': 1,
    'S-I-M2': 2,
    'F-I-M1-1': 3,
    'F-I-M1-2': 4,
}


def qualitative(percent):
    if percent >= 90:
        return 'Excelente'
    if percent >= 75:
        return 'Alto'
    if percent >= 60:
        return 'Básico'
    return 'Bajo'


wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
main = wb['Main']
eva = wb['EVA']

students = {}
for row in main.iter_rows(min_row=2, values_only=True):
    sid = row[0]
    if sid is None:
        continue
    sid = int(sid)
    students[sid] = {
        'codigo': sid,
        'nombre': row[1] or '',
        'apellido': row[2] or '',
        'curso': row[3] or '',
    }

records = {}
for row in eva.iter_rows(min_row=2, values_only=True):
    sid, puntaje, prueba, resultado = row[:4]
    if sid is None or puntaje is None or not prueba:
        continue
    sid = int(sid)
    student = students.get(sid)
    if not student or student.get('curso') != '10A':
        continue

    prueba = str(prueba).strip()
    meta = PRUEBA_META.get(prueba, {'label': prueba, 'maximo': None})
    maximo = meta['maximo']
    percent = round((puntaje / maximo) * 100) if maximo else None

    rec = records.setdefault(sid, {
        'codigo': sid,
        'nombreCompleto': f"{student['nombre']} {student['apellido']}".strip(),
        'curso': student['curso'],
        'resultados': [],
    })
    rec['resultados'].append({
        'prueba': prueba,
        'titulo': meta['label'],
        'puntaje': puntaje,
        'maximo': maximo,
        'porcentaje': percent,
        'desempeno': qualitative(percent) if percent is not None else None,
        'resultado': resultado,
    })

for rec in records.values():
    rec['resultados'].sort(key=lambda x: (ORDER.get(x['prueba'], 999), x['puntaje']))

    counts = defaultdict(int)
    totals = defaultdict(int)
    for item in rec['resultados']:
        totals[item['prueba']] += 1
    for item in rec['resultados']:
        counts[item['prueba']] += 1
        item['intento'] = counts[item['prueba']]
        item['totalIntentosPrueba'] = totals[item['prueba']]
        item['tituloVisible'] = item['titulo'] if totals[item['prueba']] == 1 else f"{item['titulo']} · Intento {counts[item['prueba']]}"

    with_pct = [r['porcentaje'] for r in rec['resultados'] if r['porcentaje'] is not None]
    rec['promedioPorcentaje'] = round(sum(with_pct) / len(with_pct)) if with_pct else None
    rec['desempenoGeneral'] = qualitative(rec['promedioPorcentaje']) if rec['promedioPorcentaje'] is not None else None
    rec['totalPruebas'] = len(rec['resultados'])

payload = {
    'curso': '10A',
    'fuente': str(WORKBOOK),
    'totalEstudiantes': len(records),
    'pruebasDisponibles': sorted({r['prueba'] for rec in records.values() for r in rec['resultados']}),
    'estudiantes': sorted(records.values(), key=lambda x: x['codigo']),
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Wrote {OUT} with {len(records)} students')
